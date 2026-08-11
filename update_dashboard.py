"""
乐博经营数据看板 - 本地更新推送脚本
功能: 读取最近7天指标.xlsx → 生成 data.js/data.json → git commit & push

用法:
    python update_dashboard.py

可被每日06:00自动化任务在取数完成后调用。

依赖: openpyxl, git (命令行)
"""
import openpyxl
import json
import subprocess
import sys
import shutil
from pathlib import Path
from datetime import datetime, timedelta

# ===== 配置 =====
WORK_DIR = Path(r"D:\龙虾工作空间")          # 指标.xlsx 所在目录
DASHBOARD_DIR = Path(__file__).parent       # 看板项目目录（脚本所在目录）
HISTORY_DAYS = 7  # 保留多少天趋势数据

# ===== Git 路径自动检测 =====
def find_git():
    """查找 git 可执行文件路径"""
    # 1. 尝试 PATH 中的 git
    git_path = shutil.which("git")
    if git_path:
        return git_path
    # 2. 尝试常见安装路径
    candidates = [
        r"C:\Program Files\Git\cmd\git.exe",
        r"C:\Program Files (x86)\Git\cmd\git.exe",
        str(Path.home() / "AppData" / "Local" / "Programs" / "Git" / "cmd" / "git.exe"),
    ]
    for p in candidates:
        if Path(p).exists():
            return p
    return None

GIT = find_git()


def get_recent_dates(days=7):
    """获取最近N天的日期列表 (YYYY-MM-DD)"""
    today = datetime.now()
    dates = []
    for i in range(days - 1, -1, -1):
        d = today - timedelta(days=i)
        dates.append(d.strftime("%Y-%m-%d"))
    return dates


def read_excel_indicators(date_str):
    """读取某天的指标.xlsx"""
    fpath = WORK_DIR / date_str / "指标.xlsx"
    if not fpath.exists():
        return None
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    indicators = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # 列结构: A=None, B=指标名称, C=路径, D=数据, E=正确数据
        name = str(row[1]) if row[1] else ""
        value = row[3] if len(row) > 3 else None
        if name:
            indicators[name] = value
    return indicators


def build_data():
    """构建完整数据结构"""
    dates = get_recent_dates(HISTORY_DAYS)
    daily_data = []

    for d in dates:
        indicators = read_excel_indicators(d)
        if indicators is not None:
            daily_data.append({"date": d, "data": indicators})
            print(f"  读取: {d} ({len(indicators)}项)")
        else:
            print(f"  跳过(无文件): {d}")

    if not daily_data:
        print("错误: 没有找到任何数据文件")
        return None

    output = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "latest_date": daily_data[-1]["date"],
        "latest": daily_data[-1]["data"],
        "history": daily_data
    }
    return output


def write_data_files(data):
    """写入 data.json 和 data.js"""
    # data.json (备份/标准格式)
    json_path = DASHBOARD_DIR / "data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  已写入: {json_path}")

    # data.js (HTML看板实际加载的文件)
    js_path = DASHBOARD_DIR / "data.js"
    js_content = (
        "// 自动生成 - 请勿手动编辑\n"
        f"// 更新时间: {data['updated_at']}\n"
        "window.__DASHBOARD_DATA__ = "
        + json.dumps(data, ensure_ascii=False, indent=2)
        + ";\n"
    )
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js_content)
    print(f"  已写入: {js_path}")


def git_push():
    """git add → commit → push"""
    print("\n=== Git 提交推送 ===")
    if not GIT:
        print("  跳过: Git 未安装或未找到")
        return False
    try:
        git = GIT  # git 可执行文件路径
        # git add
        subprocess.run([git, "add", "-A"], cwd=str(DASHBOARD_DIR), check=True)
        print("  git add 完成")

        # 检查是否有变更
        status = subprocess.run(
            [git, "status", "--porcelain"],
            cwd=str(DASHBOARD_DIR),
            capture_output=True,
            text=True
        )
        if not status.stdout.strip():
            print("  无变更，跳过提交")
            return True

        # git commit
        commit_msg = f"数据更新: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        subprocess.run(
            [git, "commit", "-m", commit_msg],
            cwd=str(DASHBOARD_DIR),
            check=True
        )
        print(f"  git commit: {commit_msg}")

        # git push
        result = subprocess.run(
            [git, "push"],
            cwd=str(DASHBOARD_DIR),
            capture_output=True,
            text=True
        )
        if result.returncode == 0:
            print("  git push 成功!")
            return True
        else:
            print(f"  git push 失败: {result.stderr}")
            return False

    except subprocess.CalledProcessError as e:
        print(f"  Git操作出错: {e}")
        return False
    except FileNotFoundError:
        print("  错误: git 命令未找到，请确保已安装 Git")
        return False


def main():
    print("=" * 50)
    print("乐博经营数据看板 - 数据更新")
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)

    # 1. 构建数据
    print("\n=== 读取指标数据 ===")
    data = build_data()
    if data is None:
        sys.exit(1)

    print(f"\n最新数据日期: {data['latest_date']}")
    print(f"历史记录: {len(data['history'])}天")

    # 2. 写入文件
    print("\n=== 写入数据文件 ===")
    write_data_files(data)

    # 3. Git 推送
    git_push()

    print("\n" + "=" * 50)
    print("完成! 看板将在 GitHub Pages 上自动更新。")
    print("=" * 50)


if __name__ == "__main__":
    main()
